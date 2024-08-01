import openpyxl
from RPA.Browser.Selenium import Selenium
import time
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl import Workbook
import requests
from io import BytesIO
from PIL import Image
import os

from utils import get_number_or_quit, get_user_dates
from staticvalue import *
from xpath import *
from logging_file import logging


class Bot:
    def __init__(self):
        logging.info("Initializing Bot class.")
        self.browser = Selenium()
        self.numbers = []
        self.browser.open_available_browser(browser_url)
        logging.info("Browser opened.")
        self.workbook = Workbook()
        self.data_sheet = self.workbook.active
        self.data_sheet.title = data_sheet
        self.data_sheet.append(header_data_sheet)
        self.image_sheet = self.workbook.create_sheet(title=image_sheet)
        self.image_count = 0
        logging.info("Bot initialized.")

    def search(self):
        logging.info("Starting search.")
        search_input_news_nyt = "Pakistan Political news"
        search_button_selector = search_button_selector_xpath
        time.sleep(5)  # Waiting for the page to load
        self.browser.click_button_when_visible(search_button_selector)
        search_input_selector = search_input_selector_xpath
        self.browser.input_text(search_input_selector, search_input_news_nyt)
        submit_button_selector = submit_button_selector_xpath
        self.browser.click_button_when_visible(submit_button_selector)
        logging.info("Search completed.")

    def select_date_range(self):
        logging.info("Selecting date range.")
        date_range_dropdown_selector = date_range_dropdown_selector_xpath
        try:
            self.browser.wait_until_element_is_visible(date_range_dropdown_selector, timeout=15)
            self.browser.click_button_when_visible(date_range_dropdown_selector)
            logging.info("Date range dropdown opened.")
        except AssertionError:
            logging.error("Date Range dropdown not found or not visible.")
            return

        date_range_options_locator = date_range_options_locator_xpath
        date_range_options = self.browser.find_elements(date_range_options_locator)
        date_range_names = [option.text for option in date_range_options]

        logging.info(f"Available Date Range options: {date_range_names}")
        selection = 4
        if 1 <= selection <= len(date_range_options):
            self.browser.click_element(date_range_options[selection - 1])
            selected_option_text = date_range_names[selection - 1]
            logging.info(f"Selected date range: {selected_option_text}")
        else:
            logging.error("Invalid selection. Please enter a number from the list.")

        if "Specific Dates" in selected_option_text:
            start_date, end_date = get_user_dates(self)
            start_date_input_selector = start_date_input_selector_xpath
            self.browser.input_text(start_date_input_selector, start_date)
            end_date_input_selector = end_date_input_selector_xpath
            self.browser.input_text(end_date_input_selector, end_date)
            end_date_input_element = self.browser.find_element(end_date_input_selector)
            end_date_input_element.send_keys(Keys.RETURN)
            logging.info(f"Selected specific date range from {start_date} to {end_date}.")

    def select_category(self, label_text, list_locator):
        logging.info(f"Selecting category: {label_text}")
        self.browser.click_element_when_visible(f"//label[text()='{label_text}']")
        categories = self.browser.find_elements(list_locator)
        category_names = [cat.text for cat in categories]

        logging.info(f"Available {label_text} categories: {category_names}")
        # category_option_list = get_number_or_quit(self)
        category_option_list = [1]
        for index in category_option_list:
            if 1 <= index <= len(categories):
                self.browser.click_element(categories[index - 1])

                logging.info(f"Selected category: {category_names[index - 1]}")
            else:
                logging.error(f"Invalid category index: {index}")

    def select_section_category(self):
        logging.info("Selecting section category.")
        self.select_category('Type', category_type_selector_xpath)

    def select_type_category(self):
        logging.info("Selecting type category.")
        self.select_category('Section', select_category_section_xpath)

    def click_show_more(self):
        logging.info("Clicking 'Show More' button.")
        show_more_button_selector = show_more_button_selector_xpath
        while True:
            try:
                self.browser.wait_until_element_is_visible(show_more_button_selector, timeout=15)
                self.browser.click_element(show_more_button_selector)
                logging.info("Clicked 'Show More' button.")
                time.sleep(2)
            except AssertionError:
                logging.info("No more 'Show More' button found or it's not visible.")
                break

    def select_sorting(self):
        logging.info("Selecting sorting option.")
        sort_by_selector = sort_by_selector_xpath
        sorting_options = {
            "1": "best",
            "2": "newest",
            "3": "oldest"
        }
        # choice = input('Enter choice Number:')
        choice = "2"  # This can be dynamic or user input
        if choice in sorting_options:
            self.browser.select_from_list_by_value(sort_by_selector, sorting_options[choice])
            logging.info(f"Selected {sorting_options[choice]} sorting.")
        else:
            logging.error("Invalid selection. Please enter a number from the list.")

    def save_search_results(self):
        logging.info("Saving search results.")
        page_source = self.browser.get_source()
        soup = BeautifulSoup(page_source, 'html.parser')

        search_results_container = soup.find("ol", {"data-testid": "search-results"})
        li_elements = search_results_container.find_all("li", class_="css-1l4w6pd")

        for index, li in enumerate(li_elements, start=1):
            date = li.find('span', class_='css-17ubb9w').text if li.find('span', class_='css-17ubb9w') else "N/A"
            region = li.find('p', class_='css-myxawk').text if li.find('p', class_='css-myxawk') else "N/A"
            title = li.find('h4', class_='css-nsjm9t').text if li.find('h4', class_='css-nsjm9t') else "N/A"
            description = li.find('p', class_='css-16nhkrn').text if li.find('p', class_='css-16nhkrn') else "N/A"
            authors = li.find('p', class_='css-1engk30').text if li.find('p', class_='css-1engk30') else "N/A"
            image_url = li.find('img', class_='css-rq4mmj')['src'] if li.find('img', class_='css-rq4mmj') else "N/A"

            logging.info(f"Result {index}: {title}, Date: {date}, Region: {region}")
            output_path = files_output_path
            os.makedirs(output_path, exist_ok=True)

            self.data_sheet.append([date, region, title, description, authors, image_url])
            if image_url != "N/A":
                self.save_image(image_url, index)

        self.workbook.save(workbook_path)
        logging.info("Search results saved.")

    def save_image(self, image_url, index):
        try:
            response = requests.get(image_url)
            response.raise_for_status()
            img = Image.open(BytesIO(response.content))
            img_path = f'{images_path}_{index}.png'
            img.save(img_path)

            img_cell = self.image_sheet.cell(row=index, column=1)
            img_cell.value = f"Image {index}"
            img_anchor = openpyxl.drawing.image.Image(img_path)
            self.image_sheet.add_image(img_anchor, f"B{index}")
            logging.info(f"Saved image {index}.")
        except Exception as e:
            logging.error(f"Error downloading or saving image {index}: {e}")
